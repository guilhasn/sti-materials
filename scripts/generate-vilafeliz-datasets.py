# -*- coding: utf-8 -*-
"""
Gerador de datasets sintéticos Vila Feliz para o módulo Power BI.
- atendimento-vila-feliz.xlsx (atendimento ao munícipe ~250 registos)
- eventos-vila-feliz.xlsx (eventos culturais ~30 registos + dimensão Espaço + dimensão Artista)

Todos os dados são fictícios. Vila Feliz é um município imaginário usado em todos os módulos
da UC Sistemas e Tecnologias de Informação (MAP, ESTG-IPL).
"""

import os
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)  # reprodutibilidade

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'docs', 'powerbi', 'datasets')
os.makedirs(OUT_DIR, exist_ok=True)


def style_header(ws, n_cols):
    """Aplica estilo ao cabeçalho (1.ª linha)."""
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    centre = Alignment(horizontal="center")
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.fill = fill
        c.alignment = centre


def autosize(ws):
    """Ajusta largura das colunas ao conteúdo."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = str(cell.value) if cell.value is not None else ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


# ==========================================================================
# DATASET 1 — ATENDIMENTO AO MUNÍCIPE
# ==========================================================================

freguesias = [
    "Vila Feliz (centro)", "Casalinho", "Carvalhais", "Pinhal Verde",
    "Ribeira dos Pássaros", "Monte Alto", "Vale Florido", "Souto da Areia",
]

departamentos = [
    "Atendimento Geral", "Urbanismo", "Águas e Saneamento",
    "Acção Social", "Educação", "Espaços Verdes", "Recursos Humanos",
    "Tesouraria",
]

canais = [
    ("Presencial", 0.55),
    ("Telefone", 0.20),
    ("Online", 0.20),
    ("Email", 0.05),
]

tipos_pedido = [
    "Pedido de informação",
    "Reclamação",
    "Licenciamento",
    "Pagamento",
    "Marcação",
    "Certidão",
    "Apoio social",
    "Outro",
]

estados = [
    ("Resolvido", 0.70),
    ("Pendente", 0.15),
    ("Encaminhado", 0.10),
    ("Cancelado", 0.05),
]


def weighted_choice(choices):
    total = sum(w for _, w in choices)
    r = random.uniform(0, total)
    upto = 0
    for c, w in choices:
        upto += w
        if upto >= r:
            return c


wb = Workbook()
ws = wb.active
ws.title = "Atendimentos"

ws.append([
    "IdAtendimento", "Data", "Hora", "Canal", "Departamento",
    "Freguesia", "TipoPedido", "TempoMinutos", "Estado",
])

start_date = date(2025, 9, 1)
end_date = date(2026, 2, 28)
n_dias = (end_date - start_date).days + 1

idx = 1000
for d in range(n_dias):
    current = start_date + timedelta(days=d)
    if current.weekday() >= 5:  # sábado / domingo
        continue
    n_atendimentos = random.randint(8, 25)
    for _ in range(n_atendimentos):
        idx += 1
        hora = f"{random.randint(9, 16):02d}:{random.randint(0, 59):02d}"
        canal = weighted_choice(canais)
        depto = random.choice(departamentos)
        freguesia = random.choice(freguesias)
        tipo = random.choice(tipos_pedido)
        # tempo varia consoante o canal
        if canal == "Presencial":
            tempo = random.randint(8, 35)
        elif canal == "Telefone":
            tempo = random.randint(3, 12)
        elif canal == "Online":
            tempo = random.randint(1, 6)
        else:
            tempo = random.randint(2, 10)
        estado = weighted_choice(estados)
        ws.append([
            f"AT-{idx:05d}", current, hora, canal, depto, freguesia,
            tipo, tempo, estado,
        ])

style_header(ws, 9)
autosize(ws)
out_path = os.path.join(OUT_DIR, "atendimento-vila-feliz.xlsx")
wb.save(out_path)
print(f"✓ {out_path}  ({ws.max_row - 1} linhas)")


# ==========================================================================
# DATASET 2 — EVENTOS CULTURAIS
# ==========================================================================

espacos = [
    ("ESP01", "Praça do Município", "Vila Feliz (centro)", "Exterior", 2000),
    ("ESP02", "Parque da Cerca", "Vila Feliz (centro)", "Exterior", 5000),
    ("ESP03", "Pavilhão Municipal", "Casalinho", "Interior", 800),
    ("ESP04", "Auditório Municipal", "Vila Feliz (centro)", "Interior", 350),
    ("ESP05", "Castelo de Vila Feliz", "Vila Feliz (centro)", "Exterior", 1500),
    ("ESP06", "Largo da Igreja", "Carvalhais", "Exterior", 600),
    ("ESP07", "Centro Cultural", "Pinhal Verde", "Interior", 250),
]

artistas = [
    ("A001", "Banda Filarmónica de Vila Feliz", "Banda", "geral@bandavf.pt"),
    ("A002", "Grupo Coral Doces Vozes", "Coral", "doces.vozes@gmail.com"),
    ("A003", "DJ Cacau", "DJ", "djcacau@booking.pt"),
    ("A004", "Companhia de Teatro Local", "Teatro", "teatrovf@local.pt"),
    ("A005", "Fadista Maria Sá", "Fado", "maria.sa@artistas.pt"),
    ("A006", "Grupo de Cantares", "Cantares", "cantares.tradicionais@vf.pt"),
    ("A007", "Bailado Júnior Vila Feliz", "Dança", "bailado.jr@vf.pt"),
    ("A008", "Tuna Académica IPL", "Tuna", "tuna@ipleiria.pt"),
    ("A009", "Banda Rock Estradas Antigas", "Rock", "estradasantigas@rock.pt"),
    ("A010", "Quarteto de Cordas Outono", "Clássica", "outono@musica.pt"),
]

tipos_evento = [
    "Festival", "Concerto", "Teatro", "Mostra Gastronómica", "Mercado",
    "Cinema ao Ar Livre", "Feira do Livro", "Exposição",
]

# Folha 1: Eventos
wb = Workbook()
ws_eventos = wb.active
ws_eventos.title = "Eventos"
ws_eventos.append([
    "codEvento", "nome", "tipo", "dataInicio", "dataFim",
    "edicao", "orcamento", "codEspaco",
])

eventos = [
    ("E001", "Festival do Chocolate", "Festival", date(2025, 10, 18), date(2025, 10, 20), "X", 35000, "ESP02"),
    ("E002", "Concerto de Outono", "Concerto", date(2025, 10, 25), date(2025, 10, 25), "VIII", 8000, "ESP04"),
    ("E003", "Mostra de Fado", "Concerto", date(2025, 11, 8), date(2025, 11, 8), "V", 6500, "ESP04"),
    ("E004", "Feira do Livro de Vila Feliz", "Feira do Livro", date(2025, 11, 15), date(2025, 11, 23), "XV", 12000, "ESP01"),
    ("E005", "Mercado Medieval", "Mercado", date(2025, 11, 22), date(2025, 11, 23), "XII", 18000, "ESP05"),
    ("E006", "Concerto de Rock", "Concerto", date(2025, 11, 29), date(2025, 11, 29), "III", 9500, "ESP03"),
    ("E007", "Cinema ao Ar Livre", "Cinema ao Ar Livre", date(2025, 12, 5), date(2025, 12, 5), "VII", 3500, "ESP02"),
    ("E008", "Iluminações de Natal", "Festival", date(2025, 12, 1), date(2026, 1, 6), "XX", 45000, "ESP01"),
    ("E009", "Concerto de Natal", "Concerto", date(2025, 12, 20), date(2025, 12, 20), "XVIII", 7800, "ESP04"),
    ("E010", "Cantares ao Menino", "Concerto", date(2025, 12, 26), date(2025, 12, 26), "X", 2500, "ESP06"),
    ("E011", "Réveillon na Praça", "Festival", date(2025, 12, 31), date(2025, 12, 31), "XV", 22000, "ESP01"),
    ("E012", "Cantar dos Reis", "Festival", date(2026, 1, 6), date(2026, 1, 6), "VI", 4500, "ESP06"),
    ("E013", "Festival de Música Clássica", "Festival", date(2026, 1, 24), date(2026, 1, 26), "IV", 16500, "ESP04"),
    ("E014", "Carnaval de Vila Feliz", "Festival", date(2026, 2, 14), date(2026, 2, 17), "XXV", 38000, "ESP01"),
    ("E015", "Mostra Gastronómica", "Mostra Gastronómica", date(2026, 2, 21), date(2026, 2, 22), "IX", 14000, "ESP02"),
    ("E016", "Teatro 'O Auto da Barca'", "Teatro", date(2026, 3, 7), date(2026, 3, 7), "II", 5500, "ESP04"),
    ("E017", "Marchas Populares", "Festival", date(2026, 6, 12), date(2026, 6, 13), "XX", 28000, "ESP01"),
]

for ev in eventos:
    ws_eventos.append(list(ev))

style_header(ws_eventos, 8)
autosize(ws_eventos)

# Folha 2: Espaços
ws_espacos = wb.create_sheet("Espacos")
ws_espacos.append(["codEspaco", "nome", "freguesia", "tipo", "lotacao"])
for esp in espacos:
    ws_espacos.append(list(esp))
style_header(ws_espacos, 5)
autosize(ws_espacos)

# Folha 3: Artistas
ws_artistas = wb.create_sheet("Artistas")
ws_artistas.append(["codArtista", "nome", "tipo", "email"])
for a in artistas:
    ws_artistas.append(list(a))
style_header(ws_artistas, 4)
autosize(ws_artistas)

# Folha 4: Actuação (relação N:M Evento <-> Artista)
ws_act = wb.create_sheet("Actuacao")
ws_act.append(["codEvento", "codArtista", "cache", "ordem"])
# associações fictícias
associations = [
    ("E001", "A001", 800, 1), ("E001", "A002", 350, 2), ("E001", "A003", 1200, 3),
    ("E002", "A010", 1500, 1),
    ("E003", "A005", 2200, 1), ("E003", "A002", 400, 2),
    ("E005", "A006", 600, 1), ("E005", "A007", 800, 2),
    ("E006", "A009", 2500, 1),
    ("E008", "A001", 800, 1), ("E008", "A007", 600, 2),
    ("E009", "A002", 500, 1), ("E009", "A001", 700, 2),
    ("E010", "A006", 400, 1),
    ("E011", "A003", 1800, 1), ("E011", "A009", 2200, 2),
    ("E013", "A010", 1800, 1), ("E013", "A005", 1500, 2),
    ("E014", "A001", 900, 1), ("E014", "A007", 700, 2), ("E014", "A008", 600, 3),
    ("E016", "A004", 3000, 1),
    ("E017", "A001", 1000, 1), ("E017", "A006", 600, 2),
]
for assoc in associations:
    ws_act.append(list(assoc))
style_header(ws_act, 4)
autosize(ws_act)

out_path = os.path.join(OUT_DIR, "eventos-vila-feliz.xlsx")
wb.save(out_path)
print(f"✓ {out_path}  ({len(eventos)} eventos, {len(espacos)} espaços, {len(artistas)} artistas, {len(associations)} actuações)")

print("\nDatasets gerados com sucesso.")
